#!/usr/bin/env python3
"""
Build the password-protected ASICS Preventa dashboard.

Reads the 'Preline order' sheet from the Excel workbook, aggregates the order by
gender / categoria / modelo (units: Compra, Extra, Cancel), renders the dashboard
template, then AES-256-GCM encrypts it with a password so the published index.html
is unreadable without that password.

Usage:
    python site/build.py --password "YOUR-PASSWORD"
    python site/build.py            # will prompt for the password

Outputs (in the repo root):
    index.html   -> public encrypted gate (the ONLY file GitHub Pages serves)
    site/app.html-> plaintext dashboard, git-ignored, kept only for local debugging
"""
import argparse
import base64
import getpass
import json
import os
import sys
import unicodedata
from pathlib import Path

import openpyxl
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes

# ---- paths ----
ROOT = Path(__file__).resolve().parent.parent
SITE = ROOT / "site"
XLSX = ROOT / "PRELINE ASICS S1-27 footwear.xlsx"
APP_TEMPLATE = SITE / "app.template.html"
APP_OUT = SITE / "app.html"          # plaintext (git-ignored)
INDEX_OUT = ROOT / "index.html"      # encrypted, published

PBKDF2_ITERS = 200_000

# Columns are located by HEADER NAME (not position) because the source layout changes
# as the workbook is edited. Keys -> list of accepted header aliases (accent/case-insensitive).
COL_ALIASES = {
    "modelo":    ["modelo2", "modelo"],
    "categoria": ["categoria", "category"],          # optional
    "gender":    ["descripcion", "genero", "gender"],
    "compra":    ["compra"],
    "extra":     ["extra"],
    "cancel":    ["cancel", "cancelado", "cancelacion"],
}
REQUIRED = ["modelo", "gender", "compra", "extra", "cancel"]  # categoria is optional


def _norm(s):
    """lowercase + strip accents + trim, for forgiving header matching."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.strip().lower()


def norm_gender(v):
    if not v:
        return "—"
    s = str(v).lower()
    if "hombre" in s:
        return "Hombre"
    if "mujer" in s:
        return "Mujer"
    return str(v)


def num(v):
    return int(v) if isinstance(v, (int, float)) else 0


def find_sheet_and_header(wb):
    """Find the sheet + header row containing the order columns. Returns (ws, header_row, colmap)."""
    for ws in wb.worksheets:
        for hr, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
            norm_cells = {_norm(c): i for i, c in enumerate(row) if c is not None}
            colmap = {}
            for key, aliases in COL_ALIASES.items():
                for a in aliases:
                    if a in norm_cells:
                        colmap[key] = norm_cells[a]
                        break
            if all(k in colmap for k in REQUIRED):
                return ws, hr, colmap
    sys.exit("ERROR: could not find a sheet with headers "
             f"{REQUIRED} (need Modelo2, Descripcion, Compra, Extra, Cancel).")


def aggregate():
    """Read the sheet -> (records, has_categoria). Columns located by header name."""
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws, header_row, col = find_sheet_and_header(wb)
    has_cat = "categoria" in col
    print(f"Sheet='{ws.title}'  header_row={header_row}  "
          f"columns={ {k: v for k, v in col.items()} }  categoria={'yes' if has_cat else 'NO'}")

    agg = {}  # (gender, cat, modelo) -> [compra, extra, cancel]
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        modelo = r[col["modelo"]]
        if not modelo:
            continue  # skip blank/separator rows
        cat = str(r[col["categoria"]]).strip() if has_cat and r[col["categoria"]] else "—"
        key = (norm_gender(r[col["gender"]]), cat, str(modelo))
        cur = agg.setdefault(key, [0, 0, 0])
        cur[0] += num(r[col["compra"]])
        cur[1] += num(r[col["extra"]])
        cur[2] += num(r[col["cancel"]])

    records = [
        {"gender": g, "categoria": c, "modelo": m,
         "compra": v[0], "extra": v[1], "cancel": v[2]}
        for (g, c, m), v in agg.items()
    ]

    totals = {k: sum(rec[k] for rec in records) for k in ("compra", "extra", "cancel")}
    if totals["compra"] <= 0:
        sys.exit(f"ERROR: Compra total is {totals['compra']} — no data parsed. Check the sheet.")
    if not has_cat:
        print("WARNING: no Categoria column found — the 'Categoría' button will be hidden.")
    print(f"OK  records={len(records)}  totals={totals}")
    return records, has_cat


def render_app(records, has_cat):
    tpl = APP_TEMPLATE.read_text(encoding="utf-8")
    payload = json.dumps(records, ensure_ascii=False, separators=(",", ":"))
    marker_start = "/*__DATA_JSON__*/"
    marker_end = "/*__END_DATA__*/"
    pre = tpl.split(marker_start)[0]
    post = tpl.split(marker_end)[1]
    html = f"{pre}{marker_start}{payload}{marker_end}{post}"
    # tell the UI whether the Categoría breakdown is available
    html = html.replace("/*__HAS_CAT__*/true/*__END_HAS_CAT__*/",
                        "true" if has_cat else "false")
    APP_OUT.write_text(html, encoding="utf-8")
    return html


def encrypt(html, password):
    salt = os.urandom(16)
    iv = os.urandom(12)
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=PBKDF2_ITERS)
    key = kdf.derive(password.encode("utf-8"))
    ct = AESGCM(key).encrypt(iv, html.encode("utf-8"), None)
    b64 = lambda b: base64.b64encode(b).decode("ascii")
    return b64(salt), b64(iv), b64(ct)


def write_gate(salt_b64, iv_b64, ct_b64):
    gate = GATE_TEMPLATE.format(
        salt=salt_b64, iv=iv_b64, ct=ct_b64, iters=PBKDF2_ITERS
    )
    INDEX_OUT.write_text(gate, encoding="utf-8")


# Public gate page: contains NO order data, only ciphertext + a Web Crypto decryptor.
GATE_TEMPLATE = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Preventa ASICS S1-27</title>
<style>
  html,body{{margin:0;height:100%;font-family:-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;}}
  body{{display:flex;align-items:center;justify-content:center;background:linear-gradient(135deg,#0b2c6f,#1763b6);}}
  .gate{{background:#fff;padding:34px 30px;border-radius:16px;box-shadow:0 18px 50px rgba(0,0,0,.3);width:330px;max-width:90vw;text-align:center;}}
  h1{{margin:0 0 4px;font-size:22px;color:#0b2c6f;font-weight:800;}}
  h1 span{{color:#ffb800;}}
  p{{margin:0 0 18px;color:#6b7686;font-size:13px;}}
  input{{width:100%;padding:12px 14px;border:1px solid #d8dee8;border-radius:9px;font-size:15px;margin-bottom:12px;}}
  input:focus{{outline:none;border-color:#1763b6;box-shadow:0 0 0 3px rgba(23,99,182,.15);}}
  button{{width:100%;padding:12px;border:0;border-radius:9px;background:#1763b6;color:#fff;font-size:15px;font-weight:700;cursor:pointer;}}
  button:hover{{background:#0b2c6f;}}
  button:disabled{{opacity:.6;cursor:wait;}}
  .err{{color:#e4002b;font-size:13px;min-height:18px;margin-top:10px;}}
</style>
</head>
<body>
  <form class="gate" id="g">
    <h1>Preventa ASICS <span>S1-27</span></h1>
    <p>Ingrese la contraseña para ver el reporte</p>
    <input id="pw" type="password" autocomplete="current-password" placeholder="Contraseña" autofocus>
    <button id="go" type="submit">Entrar</button>
    <div class="err" id="err"></div>
  </form>

<script>
const SALT="{salt}", IV="{iv}", CT="{ct}", ITERS={iters};
const b64=s=>Uint8Array.from(atob(s),c=>c.charCodeAt(0));

async function unlock(pw){{
  const enc=new TextEncoder();
  const baseKey=await crypto.subtle.importKey("raw",enc.encode(pw),"PBKDF2",false,["deriveKey"]);
  const key=await crypto.subtle.deriveKey(
    {{name:"PBKDF2",salt:b64(SALT),iterations:ITERS,hash:"SHA-256"}},
    baseKey,{{name:"AES-GCM",length:256}},false,["decrypt"]);
  const ptBuf=await crypto.subtle.decrypt({{name:"AES-GCM",iv:b64(IV)}},key,b64(CT));
  return new TextDecoder().decode(ptBuf);
}}

document.getElementById("g").addEventListener("submit",async e=>{{
  e.preventDefault();
  const btn=document.getElementById("go"), err=document.getElementById("err");
  err.textContent=""; btn.disabled=true; btn.textContent="Descifrando…";
  try{{
    const html=await unlock(document.getElementById("pw").value);
    document.open(); document.write(html); document.close();
  }}catch(_){{
    err.textContent="Contraseña incorrecta";
    btn.disabled=false; btn.textContent="Entrar";
  }}
}});
</script>
</body>
</html>
"""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--password", help="dashboard password (prompted if omitted)")
    args = ap.parse_args()
    password = args.password or getpass.getpass("Dashboard password: ")
    if not password:
        sys.exit("ERROR: empty password.")

    records, has_cat = aggregate()
    html = render_app(records, has_cat)
    salt_b64, iv_b64, ct_b64 = encrypt(html, password)
    write_gate(salt_b64, iv_b64, ct_b64)
    print(f"Wrote {INDEX_OUT}  ({INDEX_OUT.stat().st_size:,} bytes)")
    print(f"Wrote {APP_OUT} (plaintext, git-ignored)")
    print("Done. Test locally:  cd site is NOT needed; serve repo root:")
    print("    python -m http.server 8000   then open http://localhost:8000/")


if __name__ == "__main__":
    main()
